import datetime
import os
from pathlib import Path

import openpyxl
import openpyxl.styles

from chatgpt import ChatGPT

base_dir = Path(__file__).parent
excel_path = base_dir / "chat_history.xlsx"

HEADER_ROW_NUMBER = 2
ROW_HEIGHT = 15


def is_chat_history_open() -> bool:
    """excel_pathが開かれているかどうかを返す"""
    if not excel_path.exists():
        return False
    try:
        with excel_path.open("r+b"):
            pass
        return False
    except IOError:
        return True


def read_workbook() -> tuple[openpyxl.Workbook, bool]:
    """
    ワークブックを読み込みそのオブジェクトを返すとともに新規作成したかどうかを返す
    :returns: ワークブックのオブジェクト, 新規作成したかどうか
    """

    # ワークブックの読み込み
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        return wb, False
    else:
        wb = openpyxl.Workbook()
        return wb, True


def create_worksheet(title: str, wb, is_new: bool):
    """
    ワークシートを作成する
    :param title: ワークシートのタイトル
    :param wb: 対象のワークブックのオブジェクト
    :param is_new: ブックを新規作成したかどうか
    :return: ワークシートのオブジェクト
    """

    title = trim_invalid_chars(title)
    if is_new:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)  # 同じ名前がある場合、末尾に数字が付与される

    wb.move_sheet(ws, offset=-len(wb.worksheets) + 1)
    wb.active = ws

    return ws


def trim_invalid_chars(sheet_name: str) -> str:
    """
    エクセルのシート名で使えない文字列を削除
    :param sheet_name: シート名
    :return: 使えない文字列を削除したシート名
    """

    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '')
    return sheet_name


def header_formatting(ws):
    """ヘッダーの書式設定を行う"""

    ws[f"A1"].font = openpyxl.styles.Font(name="Meiryo", size=11, bold=True)

    header_a, header_b = ws[f"A{HEADER_ROW_NUMBER}"], ws[f"B{HEADER_ROW_NUMBER}"]
    font_style = openpyxl.styles.Font(name="Meiryo", size=11, bold=True, color="FFFFFF")
    header_a.font, header_b.font = font_style, font_style

    # ヘッダーの色を緑色にする
    excel_green = openpyxl.styles.PatternFill(fgColor='217346')
    header_a.fill, header_b.fill = excel_green, excel_green

    # 列の幅を調整
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 168


def write_chat_history(ws, gpt):
    """
    チャットの履歴を書き込み、書式設定する
    :param ws: ワークシートオブジェクト
    :param gpt: ChatGPTオブジェクト
    """

    font_style = openpyxl.styles.Font(name="Meiryo", size=10)
    assistant_style = openpyxl.styles.PatternFill(fgColor='d9d9d9')

    # ヘッダーの書き込み
    ws["A1"].value = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    ws[f"A{HEADER_ROW_NUMBER}"].value, ws[f"B{HEADER_ROW_NUMBER}"].value = "ロール", "発言内容"

    # チャット内容の書き込み
    for i, content in enumerate(gpt.chat_history, 3):
        cell_a, cell_b = ws[f"A{i}"], ws[f"B{i}"]

        # ロールと発言内容を書き込み
        cell_a.value, cell_b.value = content["role"], content["content"]

        # セル内改行に合わせて表示を調整
        cell_b.alignment = openpyxl.styles.Alignment(wrapText=True)

        # 行の高さを調整
        adjusted_row_height = len(content["content"].split("\n")) * ROW_HEIGHT
        ws.row_dimensions[i].height = adjusted_row_height

        # 書式設定
        cell_a.font, cell_b.font = font_style, font_style
        if content["role"] == "assistant":
            cell_a.fill, cell_b.fill = assistant_style, assistant_style


def open_workbook():
    """エクセルを開く"""
    if os.name == "nt":
        # windows
        os.system(f"start {excel_path}")
    elif os.name == "posix":
        # mac
        os.system(f"open {excel_path}")


def output_excel(gpt):
    """
    chat_history.xlsx にチャットの履歴を書き込むためのエントリポイント。
    """

    wb, is_new_create_wb = read_workbook()
    ws = create_worksheet(gpt.chat_summary, wb, is_new_create_wb)

    header_formatting(ws)
    write_chat_history(ws, gpt)
    wb.save(excel_path)
    wb.close()
    open_workbook()


if __name__ == "__main__":
    chat = ChatGPT()
    chat.chat_history = [{"role": "user", "content": "こんにちは"}, {"role": "assistant", "content": "こんばんは"}]
    chat.chat_summary = "test"
    output_excel(chat)
